"""Excel read/write helpers for DMF import generation."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from app.security import validate_users_security

REQUIRED_INPUT_COLUMNS = ("UserId", "Alias", "Email", "FirstName", "LastName")


def read_users(path: Path, sheet_name: str = "Users") -> list[dict[str, str]]:
    """Read the input users workbook. Returns one dict per non-empty data row."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        rows = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration as exc:
            raise ValueError(f"Input workbook is empty: {path}") from exc

        headers = [str(h).strip() if h is not None else "" for h in header_row]
        missing = [c for c in REQUIRED_INPUT_COLUMNS if c not in headers]
        if missing:
            raise ValueError(
                f"Input sheet is missing required columns: {', '.join(missing)}. "
                f"Found: {', '.join(h for h in headers if h)}"
            )

        users: list[dict[str, str]] = []
        for row_num, row in enumerate(rows, start=2):
            record = {
                headers[i]: _cell_to_str(row[i] if i < len(row) else None)
                for i in range(len(headers))
                if headers[i]
            }
            if not any(record.get(c) for c in REQUIRED_INPUT_COLUMNS):
                continue

            missing_values = [c for c in REQUIRED_INPUT_COLUMNS if not record.get(c)]
            if missing_values:
                raise ValueError(
                    f"Row {row_num} is missing required values: {', '.join(missing_values)}"
                )
            users.append(record)

        if not users:
            raise ValueError(f"No user rows found in {path}")
        validate_users_security(users)
        return users
    finally:
        wb.close()


def write_workbook(path: Path, headers: list[str], rows: list[list[Any]]) -> None:
    """Write a simple header + data workbook (DMF-style flat sheet)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(headers)
    for row in rows:
        ws.append(["" if v is None else v for v in row])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value).strip()
